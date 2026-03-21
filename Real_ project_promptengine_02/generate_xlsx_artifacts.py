import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

def create_xlsx(filename, data, headers, sheet_name="Sheet1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            if cell_value == "Fail":
                cell.font = Font(color="FF0000", bold=True)
    
    # Column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 60)

    wb.save(filename)
    print(f"File created: {filename}")

if __name__ == "__main__":
    output_dir = r"c:\Users\admin\Documents\GitHub\AI_Testing\AITesterBlueprint2x\Real_ project_promptengine_02\outputs"
    os.makedirs(output_dir, exist_ok=True)
    
    # 30 Scenarios
    scenarios_data = [
        ["LS_01", "Verify login with valid credentials", "Positive"],
        ["LS_02", "Verify login with invalid email", "Negative"],
        ["LS_03", "Verify login with invalid password", "Negative"],
        ["LS_04", "Verify login with blank email and password", "Negative"],
        ["LS_05", "Verify login with blank email only", "Negative"],
        ["LS_06", "Verify login with blank password only", "Negative"],
        ["LS_07", "Verify email field validation for missing '@'", "Negative"],
        ["LS_08", "Verify email field validation for missing domain", "Negative"],
        ["LS_09", "Verify password visibility toggle (eye icon)", "Functional"],
        ["LS_10", "Verify 'Remember me' checkbox selection", "Functional"],
        ["LS_11", "Verify 'Forgot Password?' link redirection", "Functional"],
        ["LS_12", "Verify 'Sign in with Google' redirection", "Integration"],
        ["LS_13", "Verify 'Sign in using SSO' redirection", "Enterprise"],
        ["LS_14", "Verify 'Sign in with Passkey' redirection", "Security"],
        ["LS_15", "Verify 'Start a FREE TRIAL' link redirection", "Functional"],
        ["LS_16", "Verify login page responsiveness on mobile devices", "UI/UX"],
        ["LS_17", "Verify login page load time (should be < 2s)", "Performance"],
        ["LS_18", "Verify error message for non-registered email", "Negative"],
        ["LS_19", "Verify account lockout after 5 failed attempts (Inference)", "Security"],
        ["LS_20", "Verify session timeout after inactivity", "Security"],
        ["LS_21", "Verify navigation back from dashboard (no re-login)", "Security"],
        ["LS_22", "Verify case sensitivity of email field (should be case-insensitive)", "Functional"],
        ["LS_23", "Verify case sensitivity of password field (should be case-sensitive)", "Functional"],
        ["LS_24", "Verify login with leading/trailing spaces in email (auto-trim)", "Functional"],
        ["LS_25", "Verify keyboard navigation (Tab key) across all fields", "Accessibility"],
        ["LS_26", "Verify page title and favicon presence", "UI"],
        ["LS_27", "Verify presence of Privacy Policy & Terms of Service links", "Compliance"],
        ["LS_28", "Verify SQL injection attempt in email field", "Security"],
        ["LS_29", "Verify XSS attempt in email field", "Security"],
        ["LS_30", "Verify login over HTTPS connection", "Security"]
    ]
    scenarios_headers = ["Scenario ID", "Description", "Type"]
    create_xlsx(os.path.join(output_dir, "test_scenarios.xlsx"), scenarios_data, scenarios_headers, "Test Scenarios")
    
    # 30 Test Cases
    cases_data = [
        ["TC_01", "Valid Login - Dashboard Check", "1. Open app.vwo.com. 2. Enter valid email. 3. Enter valid password. 4. Click 'Sign in'.", "User is redirected to the VWO dashboard successfully.", "Pass"],
        ["TC_02", "Invalid Login - Wrong Password", "1. Open app.vwo.com. 2. Enter valid email. 3. Enter incorrect password. 4. Click 'Sign in'.", "Error message: 'Your email, password, IP address or location did not match'.", "Fail"],
        ["TC_03", "Invalid Login - Unregistered Email", "1. Open app.vwo.com. 2. Enter email not in system. 3. Enter any password. 4. Click 'Sign in'.", "Error message: 'Your email, password, IP address or location did not match'.", "Pass"],
        ["TC_04", "Empty Fields Check", "1. Open app.vwo.com. 2. Leave both fields blank. 3. Click 'Sign in'.", "System prevents submission or shows inline validation error.", "Pass"],
        ["TC_05", "Email Formatting - Missing @", "1. Open app.vwo.com. 2. Enter 'testexample.com'. 3. Enter password. 4. Click 'Sign in'.", "Inline validation: 'The email address is invalid'.", "Fail"],
        ["TC_06", "Password Visibility Toggle", "1. Open app.vwo.com. 2. Type password. 3. Click eye icon.", "Password text becomes visible.", "Pass"],
        ["TC_07", "Remember Me - Session Persistence", "1. Open app.vwo.com. 2. Enter valid credentials. 3. Check 'Remember me'. 4. Login. 5. Restart browser.", "User remains logged in or email is pre-filled.", "Pass"],
        ["TC_08", "Forgot Password Redirection", "1. Open app.vwo.com. 2. Click 'Forgot Password?'.", "User is taken to the recovery page.", "Pass"],
        ["TC_09", "Google Login Integration", "1. Open app.vwo.com. 2. Click 'Sign in with Google'.", "Google authentication pop-up/page appears.", "Pass"],
        ["TC_10", "SSO Redirection", "1. Open app.vwo.com. 2. Click 'Sign in using SSO'.", "SSO login field or redirection occurs.", "Pass"],
        ["TC_11", "Passkey Redirection", "1. Open app.vwo.com. 2. Click 'Sign in with Passkey'.", "Passkey authentication prompt appears.", "Pass"],
        ["TC_12", "Free Trial Link", "1. Open app.vwo.com. 2. Click 'Start a FREE TRIAL'.", "User is taken to the signup page.", "Pass"],
        ["TC_13", "Responsive - Mobile View", "1. Open app.vwo.com in mobile simulator. 2. Check layout.", "Elements align vertically; no horizontal scrolling.", "Pass"],
        ["TC_14", "Performance - Initial Load", "1. Open network tab. 2. Refresh app.vwo.com.", "Finish time is less than 2 seconds.", "Pass"],
        ["TC_15", "HTTPS Check", "1. Observe address bar.", "Lock icon present; protocol is https.", "Pass"],
        ["TC_16", "Security - SQL Injection", "1. In email field, enter \"' OR 1=1 --\". 2. Attempt login.", "Login fails; no database error exposure.", "Fail"],
        ["TC_17", "Security - XSS", "1. In email field, enter \"<script>alert(1)</script>\". 2. Attempt login.", "Script does not execute; input is sanitized.", "Pass"],
        ["TC_18", "Case Insensitive Email", "1. Enter 'EMAIL@EXAMPLE.COM' (valid account). 2. Enter password. 3. Login.", "Login is successful.", "Pass"],
        ["TC_19", "Case Sensitive Password", "1. Enter valid email. 2. Enter password in wrong case. 3. Login.", "Login fails with error message.", "Pass"],
        ["TC_20", "Leading/Trailing Spaces", "1. Enter '  test@example.com  '. 2. Enter password. 3. Login.", "Login is successful (spaces trimmed).", "Pass"],
        ["TC_21", "Accessibility - Tab Index", "1. Place cursor in Email. 2. Press Tab.", "Focus moves to Password, then 'Remember me', then 'Sign in'.", "Pass"],
        ["TC_22", "Terms of Service Link", "1. Scroll to bottom. 2. Click 'Terms of Service'.", "Terms page opens in new tab/window.", "Pass"],
        ["TC_23", "Privacy Policy Link", "1. Scroll to bottom. 2. Click 'Privacy policy'.", "Privacy page opens in new tab/window.", "Pass"],
        ["TC_24", "Browser Back Button Post-Login", "1. Log in successfully. 2. Click browser 'Back'.", "User remains on dashboard or doesn't re-enter login form session.", "Pass"],
        ["TC_25", "Session Timeout Simulation", "1. Log in. 2. Delete session cookies. 3. Refresh.", "User is redirected back to login page.", "Pass"],
        ["TC_26", "Invalid Email Domain", "1. Enter 'test@invalid@domain.com'. 2. Login.", "System shows invalid email format.", "Pass"],
        ["TC_27", "Logo Visibility", "1. Look at top of form.", "VWO logo is visible and high quality.", "Pass"],
        ["TC_28", "Keyboard Enter Key", "1. Enter credentials. 2. Press 'Enter' on keyboard.", "Login process initiates automatically.", "Pass"],
        ["TC_29", "Multi-Tab Login", "1. Open login in two tabs. 2. Login in tab 1. 3. Refresh tab 2.", "Tab 2 reflects logged-in state or redirects to dashboard.", "Pass"],
        ["TC_30", "Login Banner Check", "1. Observe right side of page.", "VWO+ABTasty announcement is clearly visible.", "Pass"]
    ]
    cases_headers = ["Case ID", "Title", "Steps", "Expected Result", "Status"]
    create_xlsx(os.path.join(output_dir, "test_cases.xlsx"), cases_data, cases_headers, "Test Cases")
